from arch14cz_backend.view.vertical_scroll_area import VerticalScrollArea

from PySide2 import (QtWidgets, QtCore, QtGui)

from openpyxl import load_workbook
from pathlib import Path
import os

def get_columns_xlsx(path):
	
	wb = load_workbook(filename = path, read_only = True)
	columns = []
	for sheet in wb.sheetnames:
		ws = wb[sheet]
		for i, cell in enumerate(list(ws.iter_rows(max_row = 1))[0]):
			value = cell.value
			if value is not None:
				value = str(value).strip()
				if value:
					columns.append(value)
		break
	return [""] + columns

class DialogImportExcel(QtWidgets.QFrame):
	
	field_names = [
		"Lab Code",
		"C-14 Activity BP",
		"C-14 Uncert. 1 Sigma",
		"C-14 Method",
		"Delta C-13",
		"C-14 Analysis Note",
		"Country",
		"Cadastre",
		"Cadastre Code",
		"District",
		"Site Name",
		"Site Coordinates",
		"Site Note",
		"Fieldwork Event AMCR ID",
		"Activity Area",
		"Feature",
		"Context Description",
		"Depth cm",
		"Context Name",
		"Relative Dating Name 1",
		"Relative Dating Name 2",
		"Sample Number",
		"Sample Note",
		"Material Name",
		"Material Note",
		"Reliability",
		"Reliability Note",
		"Source Description",
		"Source Reference",
		"Source URI",
		"Source Acquisition",
		"Submitter Name",
		"Submitter Organization",
		"Public",
	]
	
	def __init__(self, dialog, cview):
		
		QtWidgets.QFrame.__init__(self)
		
		self._dialog = dialog
		self.cview = cview
		self.field_combos = {}
		
		self.setMinimumWidth(480)
		self.setLayout(QtWidgets.QVBoxLayout())
		
		self.path_edit = QtWidgets.QLineEdit()
		self.path_edit.textChanged.connect(self.on_path_changed)
		path_button = QtWidgets.QPushButton("Browse")
		path_button.setSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
		path_button.clicked.connect(self.on_path_button)
		path_frame = QtWidgets.QWidget()
		path_frame.setLayout(QtWidgets.QHBoxLayout())
		path_frame.layout().setContentsMargins(0, 0, 0, 0)
		path_frame.layout().addWidget(QtWidgets.QLabel("Source File:"))
		path_frame.layout().addWidget(self.path_edit)
		path_frame.layout().addWidget(path_button)
		
		field_frame = QtWidgets.QFrame()
		field_frame.setLayout(QtWidgets.QFormLayout())
		field_frame.layout().setContentsMargins(0, 0, 0, 0)
		for name in self.field_names:
			self.field_combos[name] = QtWidgets.QComboBox()
			field_frame.layout().addRow(name + ":", self.field_combos[name])
		scroll_area = VerticalScrollArea(field_frame)
		
		self.layout().addWidget(path_frame)
		self.layout().addWidget(QtWidgets.QLabel("<b>Fields:</b>"))
		self.layout().addWidget(scroll_area)
	
	@QtCore.Slot()
	def on_path_changed(self):
		
		for name in self.field_combos:
			self.field_combos[name].clear()
		
		path = self.path_edit.text()
		if not path:
			return
		if not os.path.isfile(path):
			return
		ext = os.path.splitext(path)[-1].lower()
		columns = get_columns_xlsx(path)
		if len(columns) > 1:
			for i, name in enumerate(self.field_combos.keys()):
				self.field_combos[name].addItems(columns)
				if i < len(columns) - 1:
					self.field_combos[name].setCurrentIndex(i + 1)
		
		self._dialog.set_enabled(len(columns) >= len(self.field_combos))
	
	@QtCore.Slot()
	def on_path_button(self):
		
		path, format = self.cview.get_load_path(
			"Select Source File",
			"Excel 2007+ Workbook (*.xlsx)",
		)
		if path is None:
			return
		self.path_edit.setText(path)
		self.cview.set_recent_dir(path)

