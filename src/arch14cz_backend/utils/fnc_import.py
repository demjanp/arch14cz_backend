from deposit.utils.fnc_serialize import (load_user_tool, try_numeric)
import arch14cz_backend

from arch14cz_backend.utils.fnc_phasing import (update_datings)
from arch14cz_backend.utils.fnc_convert import (float_or_none)

from openpyxl import load_workbook
from datetime import datetime
from copy import deepcopy
import csv
import os

def find_or_add_obj(cmodel, cls, data):
	
	obj = cmodel.find_object_with_descriptors([cls], data)
	if obj is None:
		obj = cmodel.add_object_with_descriptors(cls, data)
	return obj

def create_schema(cmodel):
	# create database schema and add user tools
	#
	# returns cls_lookup = {class_name: Class, ...}
	
	def load_csv(path):
		
		data = []
		with open(path, "r", newline = "", encoding="utf-8") as f:
			reader = csv.reader(f, dialect = csv.excel, delimiter = ";", quoting = csv.QUOTE_ALL)
			for i, row in enumerate(reader):
				if i == 0:
					continue
				data.append(list(row))
		return data
	
	has_form = False
	has_query = False
	for data in cmodel.get_user_tools():
		if (data["typ"] == "EntryForm") and (data["label"] == "C14 Form"):
			has_form = True
		if (data["typ"] == "Query") and (data["label"] == "Relative Datings - General"):
			has_query = True
	folder = os.path.dirname(arch14cz_backend.__file__)
	if not has_form:
		path = os.path.join(folder, "data", "c14_entry_form.txt")
		data = load_user_tool(path)
		cmodel.add_user_tool(data)
	if not has_query:
		path = os.path.join(folder, "data", "c14_query.txt")
		data = load_user_tool(path)
		cmodel.add_user_tool(data)
	
	descriptors = {
		"C_14_Analysis":	[
			"Arch14CZ_ID",
			"Lab_Code",
			"C_14_Activity_BP",
			"C_14_Uncertainty_1Sig",
			"CE_From",
			"CE_To",
			"Delta_C_13_per_mil",
			"Public",
			"Note_Analysis",
			"Note_Material",
			"Note_Sample",
			"Note_Reliability",
		],
		"Site":				["Name", "Location", "Note", "AMCR_ID"],
		"Context":			["Name", "Description", "Depth"],
		"Source":			["Description", "Reference", "URI", "Database"],
		"Activity_Area":	["Name", "AMCR_ID"],
		"Feature":			["Name", "AMCR_ID"],
		"Sample":			["Number"],
		"Material":			["Name", "AMCR_ID"],
		"Reliability":		["Name"],
		"C_14_Method":		["Name"],
		"Country":			["Name", "Code"],
		"District":			["Name", "Code"],
		"Cadastre":			["Name", "Code"],
		"Relative_Dating":	["Name", "General", "AMCR_ID"],
		"Submitter":		["Name", "Organization"],
	}
	
	cls_lookup = {}
	for cls_name in list(descriptors.keys()):
		cls_lookup[cls_name] = cmodel.add_class(cls_name)
	
	for cls_name in descriptors:
		for name in descriptors[cls_name]:
			cls_lookup[cls_name].set_descriptor(name)
	
	cls_lookup["Source"].add_relation("C_14_Analysis", "describes")
	cls_lookup["C_14_Analysis"].add_relation("Reliability", "descr")
	cls_lookup["C_14_Analysis"].add_relation("C_14_Method", "descr")
	cls_lookup["C_14_Analysis"].add_relation("Sample", "analyses")
	cls_lookup["C_14_Analysis"].add_relation("Submitter", "descr")
	cls_lookup["Sample"].add_relation("Material", "descr")
	cls_lookup["Context"].add_relation("Sample", "contains")
	cls_lookup["Context"].add_relation("Feature", "descr")
	cls_lookup["Site"].add_relation("Context", "contains")
	cls_lookup["Context"].add_relation("Activity_Area", "descr")
	cls_lookup["Cadastre"].add_relation("Site", "contains")
	cls_lookup["District"].add_relation("Cadastre", "contains")
	cls_lookup["Country"].add_relation("District", "contains")
	cls_lookup["Relative_Dating"].add_relation("Context", "dates")
	
	query = cmodel.get_query(
		"SELECT Country.Name, District.Name, Cadastre.Name, Cadastre.Code",
		silent = True
	)
	existing = set()
	for row in query:
		country, district, cadastre, code = row
		code = try_numeric(code)
		existing.add((country, district, cadastre, code))
	
	data = load_csv(os.path.join(folder, "data", "district_cadastre.csv"))
	lookup_country = {}
	lookup_district = {}
	lookup_cadastre = {}
	for country, country_code, district, district_code, cadastre, cadastre_code in data:
		country = country.strip()
		country_code = country_code.strip()
		district = district.strip()
		district_code = int(district_code.strip())
		cadastre = cadastre.strip()
		cadastre_code = int(cadastre_code.strip())
		if (country, district, cadastre, cadastre_code) in existing:
			continue
		if country not in lookup_country:
			lookup_country[country] = cmodel.add_object_with_descriptors(cls_lookup["Country"], {
				"Name": country,
				"Code": country_code,
			})
		if district not in lookup_district:
			lookup_district[district] = cmodel.add_object_with_descriptors(cls_lookup["District"], {
				"Name": district,
				"Code": district_code,
			})
		if cadastre not in lookup_cadastre:
			lookup_cadastre[cadastre] = cmodel.add_object_with_descriptors(cls_lookup["Cadastre"], {
				"Name": cadastre,
				"Code": cadastre_code,
			})
		
	for country, country_code, district, district_code, cadastre, cadastre_code in data:
		country = country.strip()
		district = district.strip()
		cadastre = cadastre.strip()
		if (country, district, cadastre, cadastre_code) in existing:
			continue
		lookup_district[district].add_relation(lookup_cadastre[cadastre], "contains")
		lookup_country[country].add_relation(lookup_district[district], "contains")
	
	data = load_csv(os.path.join(folder, "data", "relative_dating.csv"))
	for name, amcr_id in data:
		name = name.strip()
		amcr_id = amcr_id.strip()
		if amcr_id == "":
			amcr_id = None
		obj = find_or_add_obj(cmodel, cls_lookup["Relative_Dating"], {
			"Name": name
		})
		obj.set_descriptor("General", 1)
		obj.set_descriptor("AMCR_ID", amcr_id)
	
	for fname, class_name in [
		("material.csv", "Material"),
		("feature.csv", "Feature"),
		("activity_area.csv", "Activity_Area"),
	]:
		data = load_csv(os.path.join(folder, "data", fname))
		for name, amcr_id in data:
			name = name.strip()
			amcr_id = amcr_id.strip()
			if amcr_id == "":
				amcr_id = None
			find_or_add_obj(cmodel, cls_lookup[class_name], {
				"Name": name,
				"AMCR_ID": amcr_id,
			})
		
	return cls_lookup

def generate_ids(cmodel):
	
	n_rows = 0
	arch_ids = set([])
	cls_c14 = cmodel.get_class("C_14_Analysis")
	
	if cls_c14 is None:
		return n_rows, ["C_14_Analysis Class not found"]
	
	for obj in cls_c14.get_members(direct_only = True):
		n_rows += 1
		arch_id = obj.get_descriptor("Arch14CZ_ID")
		if arch_id is not None:
			arch_ids.add(arch_id)
	datestr = datetime.now().strftime("%Y%m%d")
	objs = sorted(cls_c14.get_members(direct_only = True), key = lambda obj: obj.id)
	for obj in objs:
		arch_id = obj.get_descriptor("Arch14CZ_ID")
		if arch_id is None:
			n = 1
			while True:
				arch_id = "A14CZ_%s_%04d" % (datestr, n)
				if arch_id not in arch_ids:
					break
				n += 1
			obj.set_descriptor("Arch14CZ_ID", arch_id)
			arch_ids.add(arch_id)
	
	return n_rows, []

def import_xlsx(cmodel, path, fields, progress):
	# fields[name] = column index
	#
	# returns n_imported, n_rows, errors
	#	n_imported = number of rows imported
	#	n_rows = number of rows processed
	#	errors = [error message, ...]
	
	def get_from_field(name, fields, row):
		
		index = fields.get(name, -1)
		if index == -1:
			return ""
		value = row[index].value
		if value is None:
			return ""
		return str(value).strip()
	
	errors = []
	
	c14_data = []
	
	other_data = {
		"Site": [],				# [{Name, Location, Note, AMCR_ID}, ...]
		"Context": [],			# [{Name, Description, Depth}, ...]
		"Sample": [],			# [{Number}, ...]
		"Reliability": [],		# [{Name}, ...]
		"C_14_Method": [],		# [{Name}, ...]
		"Submitter": [],		# [{Name, Organization}, ...]
	}
	
	cmodel._model.blockSignals(True)
	
	cls_lookup = create_schema(cmodel)
	
	valid_cadastres = {}
	query = cmodel.get_query(
		"SELECT Country.Name, District.Name, Cadastre.Name, Cadastre.Code",
		silent = True
	)
	for row in range(len(query)):
		country_id, country = query[row, "Country", "Name"]
		district_id, district = query[row, "District", "Name"]
		cadastre_id, cadastre = query[row, "Cadastre", "Name"]
		_, code = query[row, "Cadastre", "Code"]
		valid_cadastres[(country, district, cadastre, code)] = (
			cmodel.get_object(country_id),
			cmodel.get_object(district_id),
			cmodel.get_object(cadastre_id),
		)
	
	valid_datings = set()
	for obj in cls_lookup["Relative_Dating"].get_members(direct_only = True):
		if obj.get_descriptor("General") == 1:
			valid_datings.add(obj.get_descriptor("Name"))
	
	valid_activity_areas = {}
	for obj in cls_lookup["Activity_Area"].get_members(direct_only = True):
		valid_activity_areas[obj.get_descriptor("Name")] = obj
	
	valid_features = {}
	for obj in cls_lookup["Feature"].get_members(direct_only = True):
		valid_features[obj.get_descriptor("Name")] = obj
	
	valid_materials = {}
	for obj in cls_lookup["Material"].get_members(direct_only = True):
		valid_materials[obj.get_descriptor("Name")] = obj
	
	relative_datings = []
	sources = []
	check_site_cadastre = {}	# {site_idx: cadastre_obj_id, ...}
	check_context_site = {}		# {context_idx: site_idx, ...}
	check_sample_context = {}	# {sample_idx: context_idx, ...}
	
	wb = load_workbook(filename = path, read_only = True)
	for sheet in wb.sheetnames:
		ws = wb[sheet]
		break
	n_rows = ws.max_row - 1
	unknown_sample_n = 1
	row_n = 1
	progress.update_state(value = row_n, maximum = n_rows * 2)
	for row in ws.iter_rows(min_row = 2):
		
		progress.update_state(value = row_n)
		if progress.cancel_pressed():
			return 0, n_rows, ["Cancelled by user"]
		
		row_n += 1
		row_data = {}
		obj_country = None
		obj_district = None
		obj_cadastre = None
		obj_activity_area = None
		obj_feature = None
		obj_material = None
		
		c14_lab_code = get_from_field("Lab Code", fields, row)
		c14_activity = get_from_field("C-14 Activity BP", fields, row)
		c14_uncert = get_from_field("C-14 Uncert. 1 Sigma", fields, row)
		row_data["C_14_Method"] = get_from_field("C-14 Method", fields, row)
		c14_delta13c = get_from_field("Delta C-13", fields, row)
		c14_note_analysis = get_from_field("C-14 Analysis Note", fields, row)
		
		country = get_from_field("Country", fields, row)
		district = get_from_field("District", fields, row)
		cadastre = get_from_field("Cadastre", fields, row)
		cadastre_code = try_numeric(get_from_field("Cadastre Code", fields, row))
		
		key = (country, district, cadastre, cadastre_code)
		if key not in valid_cadastres:
			errors.append("Row %d: Invalid combination of Country, District, Cadastre, Cadastre Code" % (row_n))
			continue
		else:
			obj_country, obj_district, obj_cadastre = valid_cadastres[key]
		
		site_name = get_from_field("Site Name", fields, row)
		site_coordinates = get_from_field("Site Coordinates", fields, row)
		site_note = get_from_field("Site Note", fields, row)
		site_amcr_id = get_from_field("Fieldwork Event AMCR ID", fields, row)
		
		activity_area = get_from_field("Activity Area", fields, row)
		if activity_area not in valid_activity_areas:
			errors.append("Row %d: Invalid Activity Area" % (row_n))
			continue
		else:
			obj_activity_area = valid_activity_areas[activity_area]
		
		feature = get_from_field("Feature", fields, row)
		if feature not in valid_features:
			errors.append("Row %d: Invalid Feature" % (row_n))
			continue
		else:
			obj_feature = valid_features[feature]
		
		context_name = get_from_field("Context Name", fields, row)
		context_description = get_from_field("Context Description", fields, row)
		depth_cm = get_from_field("Depth cm", fields, row)
		
		relative_dating_name_1 = get_from_field("Relative Dating Name 1", fields, row)
		relative_dating_name_2 = get_from_field("Relative Dating Name 2", fields, row)
		
		row_data["Sample"] = get_from_field("Sample Number", fields, row)
		sample_note = get_from_field("Sample Note", fields, row)
		
		material = get_from_field("Material Name", fields, row)
		if material not in valid_materials:
			errors.append("Row %d: Invalid Material" % (row_n))
			continue
		else:
			obj_material = valid_materials[material]
		
		material_note = get_from_field("Material Note", fields, row)
		
		row_data["Reliability"] = get_from_field("Reliability", fields, row)
		reliability_note = get_from_field("Reliability Note", fields, row)
		
		source_description = get_from_field("Source Description", fields, row)
		source_reference = get_from_field("Source Reference", fields, row)
		source_uri = get_from_field("Source URI", fields, row)
		source_acquisition = get_from_field("Source Acquisition", fields, row)
		
		submitter_name = get_from_field("Submitter Name", fields, row)
		submitter_organization = get_from_field("Submitter Organization", fields, row)
		
		c14_public = get_from_field("Public", fields, row)
		
		c14_activity = float_or_none(c14_activity)
		c14_uncert = float_or_none(c14_uncert)
		c14_delta13c = float_or_none(c14_delta13c)
		
		if (not c14_lab_code):
			continue
		
		if c14_public:
			try:
				c14_public = int(c14_public)
			except:
				c14_public = 0
		
		relative_datings_row = []
		error_found = False
		for name in [relative_dating_name_1, relative_dating_name_2]:
			if name.strip() == "":
				continue
			name_general = name.split(",")[0].strip()
			if name_general not in valid_datings:
				error_found = True
				errors.append("Row %d: Invalid Relative Dating" % (row_n))
			else:
				relative_datings_row.append(name)
		if error_found:
			continue
		
		row_data["Site"] = {
			"Name": site_name,
			"Location": site_coordinates,
			"Note": site_note,
			"AMCR_ID": site_amcr_id,
		}
		
		if not context_name:
			context_name = "unspecified"
		row_data["Context"] = {
			"Name": "%s, %s" % (site_name, context_name),
			"Description": context_description,
			"Depth": depth_cm,
		}
		
		if row_data["Sample"] == "":
			row_data["Sample"] = "unknown %d" % (unknown_sample_n)
			unknown_sample_n += 1
		row_data["Sample"] = "%s-%s" % (site_name, row_data["Sample"])
		
		row_data["Sample"] = {"Number": row_data["Sample"]}
		row_data["Reliability"] = {"Name": row_data["Reliability"]}
		row_data["C_14_Method"] = {"Name": row_data["C_14_Method"]}
		row_data["Submitter"] = {
			"Name": submitter_name,
			"Organization": submitter_organization,
		}
		
		row_idxs = {}
		for key in row_data:
			if row_data[key] in other_data[key]:
				row_idxs[key] = other_data[key].index(row_data[key])
			else:
				other_data[key].append(deepcopy(row_data[key]))
				row_idxs[key] = len(other_data[key]) - 1
		
		# check for Sites in >1 Cadastre
		if row_idxs["Site"] not in check_site_cadastre:
			check_site_cadastre[row_idxs["Site"]] = obj_cadastre.id
		if obj_cadastre.id != check_site_cadastre[row_idxs["Site"]]:
			errors.append("Row %d: Site assigned to more than one Cadastre" % (row_n))
			continue
		
		# check for Contexts in >1 Site
		if row_idxs["Context"] not in check_context_site:
			check_context_site[row_idxs["Context"]] = row_idxs["Site"]
		if row_idxs["Site"] != check_context_site[row_idxs["Context"]]:
			errors.append("Row %d: Context assigned to more than one Sites" % (row_n))
			continue
		
		# check for Samples in >1 Context
		if row_idxs["Sample"] not in check_sample_context:
			check_sample_context[row_idxs["Sample"]] = row_idxs["Context"]
		if row_idxs["Context"] != check_sample_context[row_idxs["Sample"]]:
			errors.append("Row %d: Sample assigned to more than one Context" % (row_n))
			continue
		
		source_idxs = []
		sources_row = [
			{
				"Description": source_description,
				"Reference": source_reference,
				"URI": source_uri,
			},
			{
				"Description": source_acquisition,
				"Reference": "",
				"URI": "",
			},			
		]
		for data in sources_row:
			if data in sources:
				source_idxs.append(sources.index(data))
			else:
				sources.append(data)
				source_idxs.append(len(sources) - 1)
		
		relative_dating_idxs = []
		for name in relative_datings_row:
			if name in relative_datings:
				relative_dating_idxs.append(relative_datings.index(name))
			else:
				relative_datings.append(name)
				relative_dating_idxs.append(len(relative_datings) - 1)
		
		c_14_analysis = {
			"Lab_Code": c14_lab_code,
			"C_14_Activity_BP": c14_activity,
			"C_14_Uncertainty_1Sig": c14_uncert,
			"Delta_C_13_per_mil": c14_delta13c,
			"Note_Analysis": c14_note_analysis,
			"Note_Material": material_note,
			"Note_Sample": sample_note,
			"Note_Reliability": reliability_note,
			"Public": c14_public,
		}
		c14_data.append([
			deepcopy(c_14_analysis), 
			deepcopy(row_idxs), 
			deepcopy(relative_dating_idxs), 
			deepcopy(source_idxs),
			obj_cadastre,
			obj_activity_area,
			obj_feature,
			obj_material
		])
	
	if errors:
		return 0, n_rows, errors
	
	obj_lookup = dict([(cls_name, {}) for cls_name in other_data]) # {cls_name: {idx: Object, ...}, ...}
	obj_lookup["Relative_Dating"] = {}
	obj_lookup["Source"] = {}
	
	cls = cls_lookup["Relative_Dating"]
	for idx, name in enumerate(relative_datings):
		obj_lookup["Relative_Dating"][idx] = find_or_add_obj(cmodel, cls, {"Name": name})
	
	cls = cls_lookup["Source"]
	
	source_lookup = {}
	for obj in cls.get_members(direct_only = True):
		key = (obj.get_descriptor("Description"), obj.get_descriptor("Reference"), obj.get_descriptor("URI"))
		source_lookup[key] = obj
	
	progress.update_state(value = row_n, maximum = n_rows * 2 + len(sources))
	for idx, data in enumerate(sources):
		row_n += 1
		progress.update_state(value = row_n)
		if progress.cancel_pressed():
			cmodel._model.blockSignals(False)
			cmodel.on_changed([],[])
			return 0, n_rows, ["Cancelled by user"]
		key = (data["Description"], data["Reference"], data["URI"])
		if key in source_lookup:
			obj_lookup["Source"][idx] = source_lookup[key]
		else:
			obj_lookup["Source"][idx] = cmodel.add_object_with_descriptors(cls, data)
	
	for cls_name in other_data:
		for idx, data in enumerate(other_data[cls_name]):
			obj_lookup[cls_name][idx] = find_or_add_obj(cmodel, cls_lookup[cls_name], data)
	
	for (
			c_14_analysis, other_idxs, relative_dating_idxs, source_idxs, 
			obj_cadastre, obj_activity_area, obj_feature, obj_material,
		) in c14_data:
			
			row_n += 1
			progress.update_state(value = row_n)
			if progress.cancel_pressed():
				cmodel._model.blockSignals(False)
				cmodel.on_changed([],[])
				return 0, n_rows, ["Cancelled by user"]
			
			obj_c_14 = cmodel.add_object_with_descriptors(cls_lookup["C_14_Analysis"], c_14_analysis)
			
			obj_site = obj_lookup["Site"][other_idxs["Site"]]
			obj_context = obj_lookup["Context"][other_idxs["Context"]]
			obj_sample = obj_lookup["Sample"][other_idxs["Sample"]]
			obj_reliability = obj_lookup["Reliability"][other_idxs["Reliability"]]
			obj_method = obj_lookup["C_14_Method"][other_idxs["C_14_Method"]]
			obj_submitter = obj_lookup["Submitter"][other_idxs["Submitter"]]
			
			obj_c_14.add_relation(obj_reliability, "descr")
			obj_c_14.add_relation(obj_method, "descr")
			obj_c_14.add_relation(obj_sample, "analyses")
			obj_c_14.add_relation(obj_submitter, "descr")
			obj_sample.add_relation(obj_material, "descr")
			obj_context.add_relation(obj_sample, "contains")
			obj_context.add_relation(obj_feature, "descr")
			obj_site.add_relation(obj_context, "contains")
			obj_context.add_relation(obj_activity_area, "descr")
			obj_cadastre.add_relation(obj_site, "contains")
			
			for idx in relative_dating_idxs:
				obj_lookup["Relative_Dating"][idx].add_relation(obj_context, "dates")
			
			for idx in source_idxs:
				obj_lookup["Source"][idx].add_relation(obj_c_14, "describes")
	
	# check for Sites in >1 Cadastre
	for obj1 in cls_lookup["Site"].get_members(direct_only = True):
		n_found = 0
		for obj2, rel in obj1.get_relations():
			if rel != "~contains":
				continue
			if "Cadastre" not in obj2.get_class_names():
				continue
			n_found += 1
		if n_found > 1:
			errors.append("Site ID %d assigned to more than one Cadastre" % (obj1.id))
	
	# check for Contexts in >1 Site
	for obj1 in cls_lookup["Context"].get_members(direct_only = True):
		n_found = 0
		for obj2, rel in obj1.get_relations():
			if rel != "~contains":
				continue
			if "Site" not in obj2.get_class_names():
				continue
			n_found += 1
		if n_found > 1:
			errors.append("Context ID %d assigned to more than one Site" % (obj1.id))
	
	# check for Samples in >1 Context
	for obj1 in cls_lookup["Sample"].get_members(direct_only = True):
		n_found = 0
		for obj2, rel in obj1.get_relations():
			if rel != "~contains":
				continue
			if "Context" not in obj2.get_class_names():
				continue
			n_found += 1
		if n_found > 1:
			errors.append("Sample ID %d assigned to more than one Context" % (obj1.id))
	
	update_datings(cmodel)
	
	progress.stop()
	cmodel._model.blockSignals(False)
	cmodel.on_changed([],[])
	
	return len(c14_data), n_rows, errors
